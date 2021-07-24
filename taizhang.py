#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import openpyxl
from datetime import datetime

def folder(path,ex):
    xlfs = [x for x in os.listdir(path) if os.path.isfile(os.path.join(path, x))
            and os.path.splitext(os.path.join(path, x))[1] == ex]
    return xlfs
def zhuanhuan(sheet):
    data = []
    for row in sheet.values:
        l=list(row)[0:5]
        if isinstance(l[1],str):
            l[1]=l[1].strip()
        if l[1] is None:
            continue
        data.append(l)
    return data

def dataxl(path,lixl):
    data=[]
    for n in range(len(lixl)):
        pathxl=os.path.join(path,lixl[n])
        table=openpyxl.load_workbook(filename=pathxl)
        for sh in table:
            data.append(zhuanhuan(sh))
    return data

def dataxl2(path,lixl):
    data=[]
    for n in range(len(lixl)):
        pathxl=os.path.join(path,lixl[n])
        table=openpyxl.load_workbook(filename=pathxl)
        sh=table.active
        data.append(zhuanhuan(sh))
    return data
def xieru(path,lixl,lix2):
    pathxl = os.path.join(path, lixl[0])
    table = openpyxl.load_workbook(filename=pathxl)
    # print(table['Sheet1'])
    n=0
    for sh in table:
        for i in range(len(lix2[n])):
            sh.append(lix2[n][i])
        n=n+1
    table.save(pathxl)

today=datetime.now().strftime("%Y-%m")
path=os.path.join(os.path.abspath('.'),today)
zongpath=os.path.join(path,"zong")
zongtxl=folder(zongpath,".xlsx")
zongdata=dataxl(zongpath,zongtxl)
zongke=[]
for i in range(len(zongdata)):
    zongke.append(list())
    for j in range(len(zongdata[i])):
        zongke[i].append(zongdata[i][j][1])

l=["poyang","tianfanjie"]
l1=[]
l2=[]
l3=[]
for n in range(len(l)):
    l1.append(os.path.join(path,l[n]))
for i in range(len(l1)):
    print(l1[i])
    if not os.path.exists(l1[i]):
        os.makedirs(l1[i])
    l2.append(folder(l1[i],'.xlsx'))
for j in range(len(l)):
    l3.append(dataxl2(l1[j],l2[j]))
#l3[0]poyangl3[0][0]poyang第一张表
zongadd=[]
for x in range(len(l3)):
    zongadd.append(list())
    for x1 in range(len(l3[x])):
        for x2 in range(len(l3[x][x1])):
            if l3[x][x1][x2][1]  not in zongke[x]:
                zongadd[x].append(l3[x][x1][x2])
xieru(zongpath,zongtxl,zongadd)
